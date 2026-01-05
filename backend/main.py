"""FastAPI backend for LLM Council."""

from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import uuid
import json
import asyncio

from . import storage
from .council import run_full_council, generate_conversation_title, run_multi_round_debate

app = FastAPI(title="LLM Council API")

# Enable CORS for local development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


async def extract_text_from_file(file: UploadFile) -> str:
    """
    Extract text content from uploaded file.
    Supports: TXT, JSON, CSV, PDF, DOCX, XLSX
    """
    content = await file.read()
    filename = file.filename.lower()

    try:
        # PDF files
        if filename.endswith('.pdf'):
            from pypdf import PdfReader
            from io import BytesIO

            pdf_reader = PdfReader(BytesIO(content))
            text_parts = []
            for page_num, page in enumerate(pdf_reader.pages, 1):
                page_text = page.extract_text()
                if page_text.strip():
                    text_parts.append(f"[Page {page_num}]\n{page_text}")

            extracted_text = "\n\n".join(text_parts)
            return f"## File: {file.filename} (PDF, {len(pdf_reader.pages)} pages)\n```\n{extracted_text}\n```\n"

        # DOCX files
        elif filename.endswith('.docx'):
            from docx import Document
            from io import BytesIO

            doc = Document(BytesIO(content))
            paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
            extracted_text = "\n\n".join(paragraphs)
            return f"## File: {file.filename} (Word Document)\n```\n{extracted_text}\n```\n"

        # XLSX files
        elif filename.endswith(('.xlsx', '.xls')):
            from openpyxl import load_workbook
            from io import BytesIO

            wb = load_workbook(BytesIO(content), read_only=True)
            sheet_texts = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    # Filter out empty rows
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    if any(cell.strip() for cell in row_data):
                        rows.append(' | '.join(row_data))

                if rows:
                    sheet_text = f"### Sheet: {sheet_name}\n" + "\n".join(rows[:50])  # Limit to 50 rows
                    if len(list(sheet.iter_rows())) > 50:
                        sheet_text += f"\n... ({len(list(sheet.iter_rows())) - 50} more rows)"
                    sheet_texts.append(sheet_text)

            extracted_text = "\n\n".join(sheet_texts)
            return f"## File: {file.filename} (Excel Spreadsheet)\n```\n{extracted_text}\n```\n"

        # Text-based files (JSON, CSV, TXT, etc.)
        else:
            text = content.decode('utf-8')

            if filename.endswith('.json'):
                # Pretty print JSON
                data = json.loads(text)
                return f"## File: {file.filename} (JSON)\n```json\n{json.dumps(data, indent=2)}\n```\n"
            elif filename.endswith('.csv'):
                return f"## File: {file.filename} (CSV)\n```csv\n{text}\n```\n"
            else:
                # Plain text or other text-based formats
                return f"## File: {file.filename} (Text)\n```\n{text}\n```\n"

    except UnicodeDecodeError:
        return f"## File: {file.filename}\n[Binary file - content not extractable as text]\n"
    except Exception as e:
        return f"## File: {file.filename}\n[Error reading file: {str(e)}]\n"


class CreateConversationRequest(BaseModel):
    """Request to create a new conversation."""
    pass


class SendMessageRequest(BaseModel):
    """Request to send a message in a conversation."""
    content: str


class ConversationMetadata(BaseModel):
    """Conversation metadata for list view."""
    id: str
    created_at: str
    title: str
    message_count: int


class Conversation(BaseModel):
    """Full conversation with all messages."""
    id: str
    created_at: str
    title: str
    messages: List[Dict[str, Any]]


@app.get("/")
async def root():
    """Health check endpoint."""
    return {"status": "ok", "service": "LLM Council API"}


@app.get("/api/conversations", response_model=List[ConversationMetadata])
async def list_conversations():
    """List all conversations (metadata only)."""
    return storage.list_conversations()


@app.post("/api/conversations", response_model=Conversation)
async def create_conversation(request: CreateConversationRequest):
    """Create a new conversation."""
    conversation_id = str(uuid.uuid4())
    conversation = storage.create_conversation(conversation_id)
    return conversation


@app.get("/api/conversations/{conversation_id}", response_model=Conversation)
async def get_conversation(conversation_id: str):
    """Get a specific conversation with all its messages."""
    conversation = storage.get_conversation(conversation_id)
    if conversation is None:
        raise HTTPException(status_code=404, detail="Conversation not found")
    return conversation


@app.post("/api/conversations/{conversation_id}/message")
async def send_message(conversation_id: str, request: SendMessageRequest):
    """
    Send a message and run the 3-stage council process.
    Returns the complete response with all stages.
    """
    # Check if conversation exists
    conversation = storage.get_conversation(conversation_id)
    if conversation is None:
        raise HTTPException(status_code=404, detail="Conversation not found")

    # Check if this is the first message
    is_first_message = len(conversation["messages"]) == 0

    # Add user message
    storage.add_user_message(conversation_id, request.content)

    # If this is the first message, generate a title
    if is_first_message:
        title = await generate_conversation_title(request.content)
        storage.update_conversation_title(conversation_id, title)

    # Run the 3-stage council process
    stage1_results, stage2_results, stage3_result, metadata = await run_full_council(
        request.content
    )

    # Add assistant message with all stages
    storage.add_assistant_message(
        conversation_id,
        stage1_results,
        stage2_results,
        stage3_result
    )

    # Return the complete response with metadata
    return {
        "stage1": stage1_results,
        "stage2": stage2_results,
        "stage3": stage3_result,
        "metadata": metadata
    }


@app.post("/api/conversations/{conversation_id}/message/stream")
async def send_message_stream(
    conversation_id: str,
    content: str = Form(...),
    files: List[UploadFile] = File(default=[])
):
    """
    Send a message and stream the 3-stage council process.
    Returns Server-Sent Events as each stage completes.
    Supports file attachments for context.
    """
    # Check if conversation exists
    conversation = storage.get_conversation(conversation_id)
    if conversation is None:
        raise HTTPException(status_code=404, detail="Conversation not found")

    # Check if this is the first message
    is_first_message = len(conversation["messages"]) == 0

    # Extract text from uploaded files
    file_context = ""
    if files:
        print(f"Processing {len(files)} uploaded file(s)...")
        file_texts = []
        for file in files:
            print(f"  - Extracting text from: {file.filename}")
            text = await extract_text_from_file(file)
            file_texts.append(text)
        if file_texts:
            file_context = "\n\n## Attached Documents\n\n" + "\n\n".join(file_texts)
            print(f"  ✓ Successfully extracted text from {len(file_texts)} file(s)")

    # Combine user message with file context
    full_content = content + file_context
    if file_context:
        print(f"  ✓ Context enhanced with {len(file_context)} characters from attachments")

    async def event_generator():
        try:
            # Add user message (with file context if any)
            storage.add_user_message(conversation_id, full_content)

            # Start title generation in parallel (don't await yet)
            title_task = None
            if is_first_message:
                title_task = asyncio.create_task(generate_conversation_title(content))

            # Collect progress events in a list
            progress_events = []

            # Define progress callback for streaming updates
            async def progress_update(event):
                progress_events.append(event)

            # Run multi-round debate
            yield f"data: {json.dumps({'type': 'stage1_start'})}\n\n"

            # Import here to access debate internals
            from .roles import get_stage1_roles, get_juge_role
            from .openrouter import query_model
            import time

            debating_roles = get_stage1_roles()
            juge = get_juge_role()
            debate_history = []
            num_rounds = 1

            # Run debate with progress updates
            for round_num in range(1, num_rounds + 1):
                for role in debating_roles:
                    # Notify querying start
                    progress_data = {
                        'type': 'model_progress',
                        'data': {
                            'stage': 'debate',
                            'role_name': role.role_name,
                            'model': role.model,
                            'round': round_num,
                            'status': 'querying'
                        }
                    }
                    yield f"data: {json.dumps(progress_data)}\n\n"

                    # Build context (simplified from council.py)
                    from .council import _build_debate_context
                    messages = _build_debate_context(full_content, role, debate_history, round_num, num_rounds)

                    # Query model
                    start_time = time.time()
                    response = await query_model(role.model, messages, timeout=300.0)
                    elapsed_time = time.time() - start_time

                    if response is not None:
                        debate_history.append({
                            "round": round_num,
                            "role_id": role.role_id,
                            "role_name": role.role_name,
                            "model": role.model,
                            "message": response.get('content', ''),
                            "elapsed_time": elapsed_time
                        })

                        # Notify completion
                        progress_data = {
                            'type': 'model_progress',
                            'data': {
                                'stage': 'debate',
                                'role_name': role.role_name,
                                'model': role.model,
                                'round': round_num,
                                'status': 'complete',
                                'elapsed_time': elapsed_time
                            }
                        }
                        yield f"data: {json.dumps(progress_data)}\n\n"

            yield f"data: {json.dumps({'type': 'stage1_complete', 'data': debate_history})}\n\n"

            # Check if we have any responses before attempting synthesis
            if not debate_history:
                error_msg = (
                    "No models responded successfully. Please ensure Ollama is running "
                    "and models are available. Run 'ollama serve' to start Ollama."
                )
                yield f"data: {json.dumps({'type': 'error', 'message': error_msg})}\n\n"
                return

            # Juge synthesis
            yield f"data: {json.dumps({'type': 'stage3_start'})}\n\n"

            progress_data = {
                'type': 'model_progress',
                'data': {
                    'stage': 'synthesis',
                    'role_name': juge.role_name,
                    'model': juge.model,
                    'status': 'querying'
                }
            }
            yield f"data: {json.dumps(progress_data)}\n\n"

            from .council import _juge_synthesize_debate
            start_time = time.time()
            juge_synthesis = await _juge_synthesize_debate(full_content, debate_history, juge)
            elapsed_time = time.time() - start_time

            # Add elapsed time to synthesis result
            juge_synthesis['elapsed_time'] = elapsed_time

            progress_data = {
                'type': 'model_progress',
                'data': {
                    'stage': 'synthesis',
                    'role_name': juge.role_name,
                    'model': juge.model,
                    'status': 'complete',
                    'elapsed_time': elapsed_time
                }
            }
            yield f"data: {json.dumps(progress_data)}\n\n"
            yield f"data: {json.dumps({'type': 'stage3_complete', 'data': juge_synthesis})}\n\n"

            # Skip stage2 (no rankings in debate system)
            # Kept for frontend compatibility but send empty data
            stage2_results = []

            # Set stage3_result for storage
            stage3_result = juge_synthesis

            # Wait for title generation if it was started
            if title_task:
                title = await title_task
                storage.update_conversation_title(conversation_id, title)
                title_data = {'type': 'title_complete', 'data': {'title': title}}
                yield f"data: {json.dumps(title_data)}\n\n"

            # Save complete assistant message
            storage.add_assistant_message(
                conversation_id,
                debate_history,  # Saved as stage1 for compatibility
                stage2_results,  # Empty list
                stage3_result    # Juge synthesis
            )

            # Send completion event
            yield f"data: {json.dumps({'type': 'complete'})}\n\n"

        except Exception as e:
            # Send error event
            yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
        }
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
